# -----------------------------------------------------------------------------
# Arquivo : services/producao_import_service.py
# Objetivo: Ler a planilha-base mensal de produção BMG enviada pela tela de
#           configurações, identificar a competência, consolidar os dados por
#           equipe/mês/contrato, enriquecer com dds_teams e gravar a coleção
#           dds_producao_mensal no Firestore quando a importação for confirmada.
# -----------------------------------------------------------------------------

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import math
import re
from typing import Any, Iterable

from google.cloud import firestore
import openpyxl

from monitor.services.firestore_client import db


RAW_HEADER_ROW = 3
RAW_DATA_START_ROW = 4
TARGET_COLLECTION = "dds_producao_mensal"
TEAM_COLLECTION = "dds_teams"
IMPORT_JOBS_COLLECTION = "dds_import_jobs"

MONTHS = {
    "jan": 1,
    "fev": 2,
    "mar": 3,
    "abr": 4,
    "mai": 5,
    "jun": 6,
    "jul": 7,
    "ago": 8,
    "set": 9,
    "out": 10,
    "nov": 11,
    "dez": 12,
}

MONTH_LABELS = {
    1: "Janeiro",
    2: "Fevereiro",
    3: "Março",
    4: "Abril",
    5: "Maio",
    6: "Junho",
    7: "Julho",
    8: "Agosto",
    9: "Setembro",
    10: "Outubro",
    11: "Novembro",
    12: "Dezembro",
}

RAW_TO_METRIC = {
    "Equipes": "teamCount",
    "Com_efetivo": "commercialEffective",
    "Com_Impedido": "commercialBlocked",
    "Corte_efetivo": "cutEffective",
    "Corte_Impedido": "cutBlocked",
    "Serv_Emergenciais": "emergencyServices",
    "Total_Serviços": "totalServices",
    "Dias_trab": "workDays",
    "INR_Dia": "inrPerDay",
    "% Impedido": "blockedPercent",
    "KM": "km",
    "Total US": "totalUs",
    "Total R$": "totalValue",
    "HN (VAR_00)": "hnVar00",
    "HE (VAR_01)": "heVar01",
    "HEN (VAR_02)": "henVar02",
    "HED (VAR_03)": "hedVar03",
    "HEND (VAR_04)": "hendVar04",
    "SO (VAR_05)": "soVar05",
    "HNN (VAR_06)": "hnnVar06",
}

DERIVED_METRICS = ("inrPerDay", "blockedPercent", "billingPerDay", "hePerDay")


@dataclass
class RowRecord:
    source_row: int
    grupo: str | None
    contract: str
    team_key: str
    plate_raw: str | None
    metrics: dict[str, float | None]


def analyze_import_file(
    *,
    file_name: str,
    file_bytes: bytes,
    selected_month_number: int | None = None,
    selected_year: int | None = None,
    commit: bool = False,
    preview_limit: int = 10,
    resolutions: dict[str, str] | None = None,
) -> dict[str, Any]:
    if not file_bytes:
        raise ValueError("O arquivo enviado está vazio.")

    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    detected_year, detected_month_abbr = parse_filter_month_year(normalize_text(worksheet["A1"].value))
    detected_month_number = MONTHS[detected_month_abbr]

    month_number = selected_month_number or detected_month_number
    year = selected_year or detected_year
    validate_competencia(month_number=month_number, year=year)

    headers = [worksheet.cell(RAW_HEADER_ROW, col).value for col in range(1, worksheet.max_column + 1)]
    records, ignored_summary_rows = extract_records(worksheet, headers)
    deduped_records, dedup_issues = deduplicate_records(records)
    grouped = group_records(deduped_records)
    teams = load_teams_from_firestore()

    import_batch_id = f"producao_{year}_{month_number:02d}_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}"
    documents: list[tuple[str, dict[str, Any]]] = []
    issues: list[dict[str, Any]] = list(dedup_issues)
    team_updates: dict[str, dict[str, Any]] = {}

    for team_key, group_rows in sorted(grouped.items()):
        doc_id, payload, doc_issues = build_document(
            year=year,
            month_number=month_number,
            detected_month_abbr=detected_month_abbr,
            team_key=team_key,
            records=group_rows,
            team_data=teams.get(team_key),
            file_name=file_name,
            import_batch_id=import_batch_id,
            detected_year=detected_year,
            detected_month_number=detected_month_number,
            resolutions=resolutions,
            team_updates=team_updates,
        )
        documents.append((doc_id, payload))
        issues.extend(doc_issues)


    warnings_preview = build_warning_preview(
        issues=issues,
        detected_year=detected_year,
        detected_month_number=detected_month_number,
        selected_year=year,
        selected_month_number=month_number,
    )

    result = {
        "ok": True,
        "mode": "execute" if commit else "preview",
        "fileName": file_name,
        "collection": TARGET_COLLECTION,
        "detectedCompetencia": build_competencia_payload(detected_year, detected_month_number, source="A1"),
        "selectedCompetencia": build_selected_competencia_payload(
            year=year,
            month_number=month_number,
            detected_year=detected_year,
            detected_month_number=detected_month_number,
        ),
        "summary": {
            "sheetName": worksheet.title,
            "rawDataRows": len(records),
            "ignoredSummaryRows": ignored_summary_rows,
            "dedupedRows": len(deduped_records),
            "documentsToUpsert": len(documents),
            "issuesCount": len(issues),
            "missingTeamsCount": sum(1 for item in issues if item.get("type") == "city_missing_after_enrichment" and "nova/não encontrada" in item.get("message", "")),
        },
        "warningsPreview": warnings_preview,
        "issuesPreview": issues[:20],
        "documentsPreview": [
            {
                "docId": doc_id,
                "teamKey": payload.get("teamKey"),
                "contract": payload.get("contract"),
                "cityBase": payload.get("cityBase"),
                "teamType": payload.get("teamType"),
                "totalUs": (payload.get("metrics") or {}).get("totalUs", 0),
                "totalValue": (payload.get("metrics") or {}).get("totalValue", 0),
            }
            for doc_id, payload in documents[: max(0, preview_limit)]
        ],
    }

    if not commit:
        return result

    job_id = execute_import(
        documents, 
        result, 
        file_name=file_name, 
        import_batch_id=import_batch_id,
        team_updates=team_updates,
    )
    result["job"] = get_import_job(job_id)
    return result


def get_last_import_job() -> dict[str, Any] | None:
    jobs: list[tuple[str, dict[str, Any], datetime]] = []
    for snap in db.collection(IMPORT_JOBS_COLLECTION).stream():
        data = snap.to_dict() or {}
        if normalize_text(data.get("kind")) != "producao_bmg":
            continue
        executed_at = to_utc_datetime(data.get("executedAt")) or datetime.min.replace(tzinfo=timezone.utc)
        jobs.append((snap.id, data, executed_at))

    if not jobs:
        return None

    jobs.sort(key=lambda item: item[2], reverse=True)
    return normalize_import_job(jobs[0][0], jobs[0][1])


def get_import_job(job_id: str) -> dict[str, Any] | None:
    snap = db.collection(IMPORT_JOBS_COLLECTION).document(job_id).get()
    if not snap.exists:
        return None
    return normalize_import_job(snap.id, snap.to_dict() or {})


def execute_import(
    documents: list[tuple[str, dict[str, Any]]],
    analysis_result: dict[str, Any],
    *,
    file_name: str,
    import_batch_id: str,
    team_updates: dict[str, dict[str, Any]] | None = None,
) -> str:
    job_ref = db.collection(IMPORT_JOBS_COLLECTION).document()
    selected = analysis_result.get("selectedCompetencia") or {}
    detected = analysis_result.get("detectedCompetencia") or {}
    summary = analysis_result.get("summary") or {}

    job_ref.set(
        {
            "kind": "producao_bmg",
            "status": "running",
            "fileName": file_name,
            "importBatchId": import_batch_id,
            "selectedCompetencia": selected,
            "detectedCompetencia": detected,
            "summary": summary,
            "warningsPreview": analysis_result.get("warningsPreview") or [],
            "issuesPreview": analysis_result.get("issuesPreview") or [],
            "executedByName": "MONITOR WEB",
            "executedByDeviceModel": "DDS_TURNOS_MONITOR",
            "executedAt": firestore.SERVER_TIMESTAMP,
            "startedAt": firestore.SERVER_TIMESTAMP,
        }
    )

    for batch_items in chunked(documents, size=400):
        batch = db.batch()
        for doc_id, payload in batch_items:
            doc_ref = db.collection(TARGET_COLLECTION).document(doc_id)
            batch.set(
                doc_ref,
                {
                    **payload,
                    "updatedAt": firestore.SERVER_TIMESTAMP,
                    "updatedByName": "MONITOR WEB",
                    "updatedByDeviceModel": "DDS_TURNOS_MONITOR",
                },
                merge=True,
            )
        batch.commit()

    if team_updates:
        for team_chunk in chunked(list(team_updates.items()), size=400):
            batch = db.batch()
            for team_key, update_data in team_chunk:
                if not update_data:
                    continue
                team_ref = db.collection(TEAM_COLLECTION).document(team_key)
                batch.set(
                    team_ref,
                    {
                        **update_data,
                        "updatedAt": firestore.SERVER_TIMESTAMP,
                        "updatedByName": "MONITOR WEB (Importação BMG)",
                        "updatedByDeviceModel": "DDS_TURNOS_MONITOR",
                    },
                    merge=True,
                )
            batch.commit()

    job_ref.set(
        {
            "status": "success",
            "completedAt": firestore.SERVER_TIMESTAMP,
            "documentsUpserted": len(documents),
        },
        merge=True,
    )
    return job_ref.id


def extract_records(worksheet, headers: list[Any]) -> tuple[list[RowRecord], int]:
    records: list[RowRecord] = []
    ignored_summary_rows = 0

    for row_idx in range(RAW_DATA_START_ROW, worksheet.max_row + 1):
        row_map = {headers[col - 1]: worksheet.cell(row_idx, col).value for col in range(1, worksheet.max_column + 1)}
        team_key = normalize_team_key(row_map.get("veiculo"))
        if not team_key:
            ignored_summary_rows += 1
            continue

        contract = normalize_contract(row_map.get("contrato"))
        if not contract:
            continue

        metrics = {metric_key: safe_number(row_map.get(raw_col)) for raw_col, metric_key in RAW_TO_METRIC.items()}
        records.append(
            RowRecord(
                source_row=row_idx,
                grupo=normalize_text(row_map.get("Grupo")),
                contract=contract,
                team_key=team_key,
                plate_raw=normalize_plate(row_map.get("placa")),
                metrics=metrics,
            )
        )

    return records, ignored_summary_rows


def group_records(records: list[RowRecord]) -> dict[str, list[RowRecord]]:
    grouped: dict[str, list[RowRecord]] = defaultdict(list)
    for record in records:
        grouped[record.team_key].append(record)
    return grouped


def deduplicate_records(records: list[RowRecord]) -> tuple[list[RowRecord], list[dict[str, Any]]]:
    seen: dict[tuple[Any, ...], RowRecord] = {}
    issues: list[dict[str, Any]] = []

    for record in records:
        signature = duplicate_signature(record)
        if signature not in seen:
            seen[signature] = record
            continue

        existing = seen[signature]
        issues.append(
            {
                "type": "duplicate_signature_ignored",
                "contract": record.contract,
                "teamKey": record.team_key,
                "keptRow": existing.source_row,
                "ignoredRow": record.source_row,
                "message": f"Linha {record.source_row} ignorada por repetir os mesmos números da linha {existing.source_row} para {record.team_key}.",
            }
        )
        if not existing.plate_raw and record.plate_raw:
            existing.plate_raw = record.plate_raw

    return list(seen.values()), issues


def duplicate_signature(record: RowRecord) -> tuple[Any, ...]:
    metric_values = tuple(round_for_signature(record.metrics.get(metric_key)) for metric_key in RAW_TO_METRIC.values())
    return (record.contract, record.team_key, metric_values)


def build_document(
    *,
    year: int,
    month_number: int,
    detected_month_abbr: str,
    team_key: str,
    records: list[RowRecord],
    team_data: dict[str, Any] | None,
    file_name: str,
    import_batch_id: str,
    detected_year: int,
    detected_month_number: int,
    resolutions: dict[str, str] | None = None,
    team_updates: dict[str, dict[str, Any]] | None = None,
) -> tuple[str, dict[str, Any], list[dict[str, Any]]]:
    issues: list[dict[str, Any]] = []

    consolidated_metrics: dict[str, float] = defaultdict(float)
    observed_plates: list[str] = []
    observed_groups: list[str] = []
    observed_contracts: list[str] = []

    for record in records:
        if record.plate_raw:
            observed_plates.append(record.plate_raw)
        if record.grupo:
            observed_groups.append(record.grupo)
        if record.contract:
            observed_contracts.append(record.contract)
        for metric_key, value in record.metrics.items():
            if value is None:
                continue
            consolidated_metrics[metric_key] += value

    total_services = consolidated_metrics.get("totalServices", 0.0)
    work_days = consolidated_metrics.get("workDays", 0.0)
    commercial_effective = consolidated_metrics.get("commercialEffective", 0.0)
    commercial_blocked = consolidated_metrics.get("commercialBlocked", 0.0)
    total_value = consolidated_metrics.get("totalValue", 0.0)
    he_var01 = consolidated_metrics.get("heVar01", 0.0)

    blocked_denominator = commercial_effective + commercial_blocked
    consolidated_metrics["inrPerDay"] = (total_services / work_days) if work_days else 0.0
    consolidated_metrics["blockedPercent"] = (commercial_blocked / blocked_denominator) if blocked_denominator else 0.0
    consolidated_metrics["billingPerDay"] = (total_value / work_days) if work_days else 0.0
    consolidated_metrics["hePerDay"] = (he_var01 / work_days) if work_days else 0.0

    team_contract = normalize_contract(first_value(team_data, "contract", "Contrato", "contrato"))
    team_plate = normalize_plate(first_value(team_data, "plate", "placa", "PLACA"))
    city_base = normalize_text(first_value(team_data, "cityBase", "cidadeBase", "Cidade_Base", "Cidade", "cidade"))
    team_type = normalize_text(first_value(team_data, "teamType", "tipo", "Tipo"))
    electrician1 = normalize_text(first_value(team_data, "electrician1", "eletricista1", "Elet. 1", "Elet 1"))
    electrician2 = normalize_text(first_value(team_data, "electrician2", "eletricista2", "Elet.2", "Elet 2"))
    base = normalize_text(first_value(team_data, "base", "BASE", "Base"))
    agency = normalize_text(first_value(team_data, "agency", "agencia", "Agencia"))
    display_name = normalize_text(first_value(team_data, "displayName")) or team_key
    members = string_list(first_value(team_data, "members", default=[]))

    final_contract = ""
    distinct_contracts = sorted(set(observed_contracts))
    main_raw_contract = distinct_contracts[-1] if distinct_contracts else ""

    if team_contract and team_contract not in distinct_contracts:
        res_key = f"contract_mismatch:{team_key}"
        decision = (resolutions or {}).get(res_key)
        if decision == "keep_db":
            final_contract = team_contract
        elif decision == "use_sheet":
            final_contract = main_raw_contract
            if team_updates is not None:
                team_updates.setdefault(team_key, {})["contract"] = main_raw_contract
        else:
            final_contract = main_raw_contract
            issues.append(
                {
                    "id": res_key,
                    "type": "contract_mismatch_vs_dds_teams",
                    "teamKey": team_key,
                    "rawContract": main_raw_contract,
                    "teamContract": team_contract,
                    "message": f"Contrato{'s' if len(distinct_contracts)>1 else ''} da planilha ({'/'.join(distinct_contracts)}) diverge{'m' if len(distinct_contracts)>1 else ''} do sistema ({team_contract}).",
                }
            )
    else:
        final_contract = team_contract if team_contract else main_raw_contract

    distinct_plates = sorted({plate for plate in observed_plates if plate})
    chosen_plate = team_plate or (distinct_plates[0] if distinct_plates else None)
    if len(distinct_plates) > 1:
        res_key = f"multiple_raw_plates:{team_key}"
        issues.append(
            {
                "id": res_key,
                "type": "multiple_raw_plates",
                "teamKey": team_key,
                "contract": final_contract,
                "rawPlates": distinct_plates,
                "chosenPlate": chosen_plate,
                "message": f"Foram encontradas placas divergentes no Excel para {team_key}: {', '.join(distinct_plates)}.",
            }
        )
    if team_data is None and team_updates is not None:
        if final_contract:
            team_updates.setdefault(team_key, {})["contract"] = final_contract
        if chosen_plate:
            team_updates.setdefault(team_key, {})["plate"] = chosen_plate

    if not city_base:
        res_key = f"city_missing:{team_key}"
        decision_city_base = (resolutions or {}).get(res_key)
        if decision_city_base and decision_city_base.strip():
            city_base = decision_city_base.strip()
            if team_updates is not None:
                team_updates.setdefault(team_key, {})["cityBase"] = city_base
        else:
            is_new = team_data is None
            msg = f"Equipe {team_key} nova/não encontrada na base. Digite a Cidade/Base para cadastrá-la." if is_new else f"Cidade não encontrada no cadastro-base para {team_key}."
            issues.append(
                {
                    "id": res_key,
                    "type": "city_missing_after_enrichment",
                    "teamKey": team_key,
                    "contract": final_contract,
                    "message": msg,
                }
            )

    goal = derive_goal(final_contract, team_type)
    month_abbr = month_number_to_abbr(month_number)
    month_key = f"{year:04d}-{month_number:02d}"
    doc_id = f"{year}_{month_number:02d}_{final_contract}_{team_key}"

    metrics_payload = {
        metric_key: round(consolidated_metrics.get(metric_key, 0.0), 6)
        for metric_key in sorted(set(RAW_TO_METRIC.values()).union(DERIVED_METRICS))
    }

    payload: dict[str, Any] = {
        "teamKey": team_key,
        "displayName": display_name,
        "members": members,
        "year": year,
        "monthNumber": month_number,
        "monthKey": month_key,
        "monthAbbr": month_abbr,
        "contract": final_contract,
        "plate": chosen_plate,
        "cityBase": city_base,
        "teamType": goal.get("type"),
        "agency": agency,
        "base": base,
        "goal": goal,
        "metrics": metrics_payload,
        "source": {
            "kind": "base_bmg_raw_upload",
            "file": file_name,
            "sheet": "Sheet1",
            "sourceRows": [record.source_row for record in records],
            "mergedRowCount": len(records),
            "rawPlates": distinct_plates,
            "grupo": sorted(set(group for group in observed_groups if group)),
            "importBatchId": import_batch_id,
            "detectedCompetencia": build_competencia_payload(detected_year, detected_month_number, source="A1"),
            "selectedCompetencia": build_competencia_payload(year, month_number, source="selecionada"),
            "detectedMonthAbbr": detected_month_abbr,
            "importedAtIso": datetime.now(timezone.utc).isoformat(),
        },
        "enrichment": {
            "teamFound": bool(team_data),
            "cityResolvedFrom": "dds_teams" if city_base else None,
            "plateResolvedFrom": "dds_teams" if team_plate else ("arquivo" if chosen_plate else None),
        },
    }

    if team_data:
        payload["teamRef"] = f"{TEAM_COLLECTION}/{team_key}"

    return doc_id, payload, issues


def build_warning_preview(
    *,
    issues: list[dict[str, Any]],
    detected_year: int,
    detected_month_number: int,
    selected_year: int,
    selected_month_number: int,
) -> list[str]:
    warnings: list[str] = []
    if detected_year != selected_year or detected_month_number != selected_month_number:
        warnings.append(
            f"A competência selecionada ({month_number_to_label(selected_month_number)}/{selected_year}) difere da detectada no arquivo ({month_number_to_label(detected_month_number)}/{detected_year})."
        )

    for issue in issues[:6]:
        message = normalize_text(issue.get("message"))
        if message:
            warnings.append(message)

    return warnings[:8]


def build_competencia_payload(year: int, month_number: int, *, source: str) -> dict[str, Any]:
    return {
        "year": year,
        "monthNumber": month_number,
        "monthKey": f"{year:04d}-{month_number:02d}",
        "label": f"{month_number_to_label(month_number)}/{year}",
        "source": source,
    }


def build_selected_competencia_payload(
    *,
    year: int,
    month_number: int,
    detected_year: int,
    detected_month_number: int,
) -> dict[str, Any]:
    payload = build_competencia_payload(year, month_number, source="usuario")
    payload["matchesDetected"] = detected_year == year and detected_month_number == month_number
    return payload


def validate_competencia(*, month_number: int, year: int) -> None:
    if month_number not in MONTH_LABELS:
        raise ValueError("Informe um mês válido para a competência.")
    if year < 2020 or year > 2100:
        raise ValueError("Informe um ano válido para a competência.")


def parse_filter_month_year(filter_text: str | None) -> tuple[int, str]:
    if not filter_text:
        raise ValueError("Não foi possível localizar o texto de filtros na célula A1.")
    month_match = re.search(r"Nome do M[êe]s\s+é\s+([A-Za-zçÇãõáéíóúÁÉÍÓÚ]+)", filter_text, flags=re.IGNORECASE)
    year_match = re.search(r"Ano\s+é\s+(\d{4})", filter_text, flags=re.IGNORECASE)
    if not month_match or not year_match:
        raise ValueError("Não foi possível extrair mês e ano do conteúdo de A1.")
    month_abbr = month_match.group(1).strip().lower()[:3]
    if month_abbr not in MONTHS:
        raise ValueError(f"Mês não reconhecido em A1: {month_match.group(1)!r}")
    return int(year_match.group(1)), month_abbr


def load_teams_from_firestore() -> dict[str, dict[str, Any]]:
    teams: dict[str, dict[str, Any]] = {}
    for doc in db.collection(TEAM_COLLECTION).stream():
        data = doc.to_dict() or {}
        team_key = normalize_team_key(data.get("teamKey") or doc.id)
        if not team_key:
            continue
        data["teamKey"] = team_key
        teams[team_key] = data
    return teams


def normalize_import_job(job_id: str, data: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": job_id,
        "kind": normalize_text(data.get("kind")) or "producao_bmg",
        "status": normalize_text(data.get("status")) or "unknown",
        "fileName": normalize_text(data.get("fileName")),
        "selectedCompetencia": data.get("selectedCompetencia") or {},
        "detectedCompetencia": data.get("detectedCompetencia") or {},
        "summary": data.get("summary") or {},
        "documentsUpserted": safe_int(data.get("documentsUpserted")),
        "executedByName": normalize_text(data.get("executedByName")) or "MONITOR WEB",
        "executedAt": normalize_timestamp_string(data.get("executedAt")),
        "completedAt": normalize_timestamp_string(data.get("completedAt")),
        "warningsPreview": data.get("warningsPreview") or [],
        "issuesPreview": data.get("issuesPreview") or [],
    }


def normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def normalize_team_key(value: Any) -> str | None:
    text = normalize_text(value)
    return text.upper() if text else None


def normalize_plate(value: Any) -> str | None:
    text = normalize_text(value)
    return re.sub(r"\s+", "", text).upper() if text else None


def normalize_contract(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float):
        if math.isnan(value):
            return None
        if value.is_integer():
            return str(int(value))
    if isinstance(value, int):
        return str(value)
    text = normalize_text(value)
    if text is None:
        return None
    return text[:-2] if text.endswith(".0") else text


def to_utc_datetime(value: Any) -> datetime | None:
    if value is None:
        return None
    if hasattr(value, "to_datetime"):
        value = value.to_datetime()
    if isinstance(value, datetime):
        return value.astimezone(timezone.utc) if value.tzinfo else value.replace(tzinfo=timezone.utc)
    text = str(value).strip()
    if not text:
        return None
    try:
        parsed = datetime.fromisoformat(text.replace("Z", "+00:00"))
    except Exception:
        return None
    return parsed.astimezone(timezone.utc) if parsed.tzinfo else parsed.replace(tzinfo=timezone.utc)


def normalize_timestamp_string(value: Any) -> str | None:
    if value is None:
        return None
    dt = to_utc_datetime(value)
    if dt is not None:
        return dt.isoformat()
    text = str(value).strip()
    return text or None


def safe_int(value: Any) -> int | None:
    try:
        if value is None:
            return None
        return int(value)
    except Exception:
        return None


def safe_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return float(int(value))
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def round_for_signature(value: float | None) -> float | None:
    return round(value, 6) if value is not None else None


def chunked(items: list[tuple[str, dict[str, Any]]], size: int = 400) -> Iterable[list[tuple[str, dict[str, Any]]]]:
    for index in range(0, len(items), size):
        yield items[index:index + size]


def first_value(data: dict[str, Any] | None, *keys: str, default: Any = None) -> Any:
    if not isinstance(data, dict):
        return default
    for key in keys:
        if key in data and data.get(key) not in (None, ""):
            return data.get(key)
    return default


def derive_goal(contract: str | None, team_type: str | None) -> dict[str, Any]:
    team_type_upper = (normalize_text(team_type) or "").upper()
    contract_str = contract or ""

    if "CESTO" in team_type_upper or contract_str.endswith("25149"):
        name = "STC - Servicos Técnicos Comerciais (CESTO)"
        target = 811
    else:
        name = "STC - Servicos Técnicos Comerciais"
        target = 859

    return {"type": name, "targetUs": target}


def month_number_to_label(month_number: int) -> str:
    return MONTH_LABELS.get(month_number, f"Mês {month_number:02d}")


def month_number_to_abbr(month_number: int) -> str:
    for abbr, number in MONTHS.items():
        if number == month_number:
            return abbr
    return f"m{month_number:02d}"


def string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        out: list[str] = []
        seen: set[str] = set()
        for item in value:
            text = normalize_text(item)
            if not text:
                continue
            folded = text.casefold()
            if folded in seen:
                continue
            seen.add(folded)
            out.append(text)
        return out
    text = normalize_text(value)
    return [text] if text else []
